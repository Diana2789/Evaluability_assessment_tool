import streamlit as st
import openai
import os
import openpyxl
import pypdf
import docx

# 1. Ֆունկցիա՝ Ծրագրային փաստաթղթից (PDF կամ DOCX) տեքստը դուրս բերելու համար
def extract_text_from_file(uploaded_file):
    if uploaded_file.name.endswith('.pdf'):
        pdf_reader = pypdf.PdfReader(uploaded_file)
        text = ""
        for page in pdf_reader.pages:
            text_content = page.extract_text()
            if text_content:
                text = text + "\n" + text_content
        return text
    elif uploaded_file.name.endswith('.docx'):
        doc = docx.Document(uploaded_file)
        return "\n".join([p.text for p in doc.paragraphs])
    return None

# 2. AI Գնահատման ֆունկցիա (հարց առ հարց)
def ai_evaluate_criterion(document_text, criterion_id, criterion_title, hint_text):
    system_prompt = """
    Դուք բյուջետային ծրագրերի և միջոցառումների գնահատման փորձագետ եք: Ձեր խնդիրն է վերլուծել օգտատիրոջ կողմից ներբեռնված ծրագրային փաստաթուղթը և տալ գնահատական համաձայն տրված չափանիշի և հուշման:
    
    Խստորեն հետևեք հետևյալ կանոններին.
    1. Գնահատականը պետք է լինի միայն այս երեք տարբերակներից մեկը՝ "Այո", "Ոչ", "Կիրառելի չէ": Եթե չափանիշը բավարարված է՝ "Այո", եթե ոչ՝ "Ոչ":
    2. Տվեք մանրամասն պրոֆեսիոնալ հիմնավորում (հայերենով):
    3. Տվեք կոնկրետ լավարկման (բարելավման) առաջարկություն, թե ինչպես փոխել փաստաթուղթը, որ թերությունը վերանա:
    4. Նշեք աղբյուրները (էջ, բաժին կամ մեջբերում):

    Պատասխանը վերադարձրեք ԽԻՍՏ հետևյալ ձևաչափով (առանց ավելորդ տեքստի, օգտագործեք ||| որպես բաժանարար).
    ԳՆԱՀԱՏԱԿԱՆ|||ՀԻՄՆԱՎՈՐՈՒՄ|||ԱՌԱՋԱՐԿՈՒԹՅՈՒՆ|||ԱՂԲՅՈՒՐ
    
    Օրինակ՝
    Այո|||Ծրագրի անվանումը հստակ արտահայտում է պետության միջամտությունը, քանի որ նշված է սուբսիդավորման մեխանիզմը:|||Առաջարկություն չկա:|||Էջ 2, Բաժին 1.2
    """
    
    user_content = f"""
    ՉԱՓԱՆԻՇԻ Հ/Հ: {criterion_id}
    ՉԱՓԱՆԻՇ: {criterion_title}
    ՀՈՒՇՈՒՄ/ՄԵԹՈԴԱԲԱՆՈՒԹՅՈՒՆ: {hint_text}
    
    Ահա գնահատվող ծրագրային փաստաթղթի տեքստը: Վերլուծիր այն և տուր պատասխանը սահմանված ձևաչափով.
    
    --- ՓԱՍՏԱԹՂԹԻ ՍԿԻԶԲ ---
    {document_text}
    --- ՓԱՍՏԱԹՂԹԻ ԱՎԱՐՏ ---
    """
    
    try:
        response = openai.chat.completions.create(
            model="gpt-4o",  # Օգտագործվում է gpt-4o մոդելը՝ խորը և ճշգրիտ վերլուծության համար
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_content}
            ],
            temperature=0.1  # Ցածր ջերմաստիճան՝ կայունության համար
        )
        res_text = response.choices[0].message.content.strip()
        parts = res_text.split("|||")
        if len(parts) == 4:
            return parts[0].strip(), parts[1].strip(), parts[2].strip(), parts[3].strip()
        else:
            return "Ոչ", "Սխալ՝ AI-ի պատասխանի ձևաչափի խախտում:", "Վերանայել ձեռքով", "Չկա"
    except Exception as e:
        return "Ոչ", f"AI հարցման սխալ. {str(e)}", "Փորձել կրկին", "Չկա"

# --- STREAMLIT ՎԵԲ ԻՆՏԵՐՖԵՅՍ ---
st.set_page_config(page_title="ԳԳ Ավտոմատացված Հարթակ", page_icon="📊", layout="wide")

st.title("📊 Միջոցառումների Գնահատելիության Գնահատման (ԳԳ) Ավտոմատ Հարթակ")
st.subheader("Ներբեռնեք ծրագիրը և ստացեք լրացված պաշտոնական Excel գործիքը")

# Օգտատերը ձախ մենյուում պետք է մուտքագրի իր OpenAI API Key-ը
api_key = st.sidebar.text_input("Մուտքագրեք Ձեր OpenAI API Key-ը", type="password")
if api_key:
    openai.api_key = api_key

st.markdown("""
### Ինչպե՞ս է աշխատում համակարգը.
1. **Ներբեռնեք ԳԳ դատարկ Excel գործիքը** (այն ֆայլը, որն ունի ներդրված բանաձևերը):
2. **Ներբեռնեք Գնահատվող ծրագիրը / միջոցառումը** (PDF կամ Word ձևաչափով):
3. Համակարգը ավտոմատ կընթերցի հարցերը Excel-ից, կկատարի AI անալիզ, կլրացնի սյունակները և թույլ կտա ներբեռնել պատրաստի ֆայլը:
""")

col1, col2 = st.columns(2)

with col1:
    excel_file = st.file_uploader("1. Ներբեռնեք Ձեր ԳԳ Excel գործիքը (.xlsx)", type=["xlsx"])

with col2:
    doc_file = st.file_uploader("2. Ներբեռնեք Գնահատվող Ծրագրային Փաստաթուղթը (.pdf, .docx)", type=["pdf", "docx"])

if st.button("🚀 Սկսել Գնահատումը"):
    if not api_key:
        st.error("Խնդրում ենք ձախ կողմում մուտքագրել Ձեր OpenAI API Key-ը:")
    elif excel_file is None or doc_file is None:
        st.error("Խնդրում ենք ներբեռնել թե՛ Excel գործիքը, թե՛ ծրագրային փաստաթուղթը:")
    else:
        with st.spinner("Կարդացվում են փաստաթղթերը և կատարվում է AI վերլուծություն... Սա կարող է տևել 1-2 րոպե:"):
            
            # 1. Կարդալ ծրագրային տեքստը
            document_text = extract_text_from_file(doc_file)
            
            # 2. Բացել Excel-ը openpyxl-ով (պահպանելով բանաձևերը)
            wb = openpyxl.load_workbook(excel_file, data_only=False)
            
            # Գտնում ենք հարցաթերթի թերթիկը (Sheet 2)
            sheet_name = None
            for name in wb.sheetnames:
                if "հարցաթերթ" in name.lower() or "sheet2" in name.lower() or "2." in name.lower():
                    sheet_name = name
                    break
            
            if not sheet_name:
                sheet_name = wb.sheetnames[0] # եթե հատուկ անունը չգտնի, վերցնում է առաջին էջը
                
            sheet = wb[sheet_name]
            
            # Գտնում ենք սյունակների ինդեքսները ըստ Հեդերի
            header_row = 1
            for r in range(1, 6):
                if sheet.cell(row=r, column=1).value == "Հ/Հ" or sheet.cell(row=r, column=2).value == "ՉԱՊԱՆԻՇ":
                    header_row = r
                    break
            
            st.info(f"Աշխատում ենք Excel-ի '{sheet_name}' թերթիկի հետ (Տող {header_row}):")
            
            # Ցիկլով անցնում ենք բոլոր տողերով header-ից հետո
            evaluated_count = 0
            max_row = sheet.max_row
            
            progress_bar = st.progress(0)
            
            for row_idx in range(header_row + 1, max_row + 1):
                # Թարմացնում ենք Progress Bar-ը
                progress_bar.progress((row_idx - header_row) / (max_row - header_row))
                
                c_id = str(sheet.cell(row=row_idx, column=1).value or "").strip()
                c_title = str(sheet.cell(row=row_idx, column=2).value or "").strip()
                c_hint = str(sheet.cell(row=row_idx, column=3).value or "").strip()
                
                # Եթե տողը դատարկ չէ և ունի ենթակետի կառուցվածք (օր.՝ 1.1, 12.1 և այլն)
                if c_id and ("." in c_id or c_id.isdigit()):
                    if c_title and c_title != "None" and len(c_hint) > 5:
                        
                        # Կանչում ենք AI-ին գնահատելու
                        decision, justification, recommendation, sources = ai_evaluate_criterion(
                            document_text, c_id, c_title, c_hint
                        )
                        
                        # Լրացնում ենք Excel-ի համապատասխան բջիջները
                        sheet.cell(row=row_idx, column=4).value = decision         # D սյունակ - ԳՆԱՀԱՏԱԿԱՆ (Այո/Ոչ/Կիրառելի չէ)
                        sheet.cell(row=row_idx, column=6).value = justification    # F սյունակ - ԳՆԱՀԱՏԱԿԱՆԻ ՀԻՄՆԱՎՈՐՈՒՄ
                        sheet.cell(row=row_idx, column=7).value = recommendation   # G սյունակ - ԼԱՎԱՐԿՄԱՆ ԱՌԱՋԱՐԿՈՒԹՅՈՒՆ
                        sheet.cell(row=row_idx, column=8).value = sources          # H սյունակ - ԱՂԲՅՈՒՐՆԵՐ
                        
                        evaluated_count += 1
            
            # Պահպանում ենք արդյունքը ժամանակավոր ֆայլում
            output_filename = "Լրացված_ԳԳ_Գործիք.xlsx"
            wb.save(output_filename)
            
            st.success(f"Գնահատումն ավարտվեց։ Հաջողությամբ վերլուծվեց {evaluated_count} չափանիշ։")
            
            # Ֆայլի ներբեռնման (Download) կոճակ օգտատիրոջ համար
            with open(output_filename, "rb") as file:
                st.download_button(
                    label="📥 Ներբեռնել Լրացված Excel Ֆայլը",
                    data=file,
                    file_name=output_filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
